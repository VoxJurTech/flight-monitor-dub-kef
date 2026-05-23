"""
Geracao de relatorios:
- Excel (.xlsx) com 4 abas: Atual, Top Hubs, Historico, Divergencias
- HTML dashboard com Chart.js

Modelo: uma viagem round-trip GRU<->KEF com hub de conexao na ida.
Cada cotacao no DB e' uma opcao retornada pelo Google Flights, com hub
identificado na coluna 'hub'.
"""

import json
import logging
import sqlite3
from datetime import datetime, timezone, timedelta
from pathlib import Path

from filters import in_arrival_window, has_window, window_label

# Brasil = UTC-3 fixo (sem horario de verao desde 2019).
BRT = timezone(timedelta(hours=-3))


def _to_brt(iso_utc, fmt: str = "%d/%m/%Y %H:%M BRT") -> str:
    """Converte ISO 8601 UTC -> string BRT formatada. Vazio se input vazio."""
    if not iso_utc:
        return ""
    try:
        dt = datetime.fromisoformat(str(iso_utc))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(BRT).strftime(fmt)
    except (ValueError, TypeError):
        return str(iso_utc or "")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)


# ---------- Excel ----------

def generate_excel(storage, config: dict, output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_top_hubs(wb, storage, config)
    _sheet_current(wb, storage, config)
    _sheet_history(wb, storage, config)
    _sheet_divergences(wb, storage, config)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info(f"  Excel salvo: {output_path}")


def _sheet_top_hubs(wb, storage, config: dict):
    """Melhor preco round-trip GRU<->KEF por hub de conexao."""
    ws = wb.create_sheet("Top Hubs")
    ws.append([
        "#", "Hub", "Cidade", "Preço total (BRL)", "Companhia(s)",
        "Duração (h)", "Escalas", "Última cotação (BRT)",
    ])
    _header_style(ws, 1)

    trip = config["trip"]
    hubs_by_code = {h["code"]: h for h in config["hubs"]}

    rows = storage.get_latest_by_hub(
        trip["origin"], trip["destination"],
        trip["outbound_date"], trip["return_date"],
    )
    for i, q in enumerate(rows, 1):
        hub = q["hub"]
        city = hubs_by_code.get(hub, {}).get("city", "")
        ws.append([
            i, hub, city,
            float(q["price"]),
            q.get("airline") or "",
            round((q.get("duration_min") or 0) / 60, 1),
            q.get("transfers") or 0,
            _to_brt(q.get("fetched_at")),
        ])

    if len(rows) >= 2:
        last = ws.max_row
        rule = ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
        ws.conditional_formatting.add(f"D2:D{last}", rule)

    _autosize_columns(ws)


def _sheet_current(wb, storage, config: dict):
    """Todas as cotacoes do ultimo sweep, ordenadas por preco."""
    ws = wb.create_sheet("Atual")
    ws.append([
        "Origem", "Destino", "Hub", "Ida", "Volta", "Preço (BRL)",
        "Companhia(s)", "Duração (h)", "Escalas", "Última cotação (BRT)",
    ])
    _header_style(ws, 1)

    trip = config["trip"]
    rows = storage.get_latest_by_hub(
        trip["origin"], trip["destination"],
        trip["outbound_date"], trip["return_date"],
    )
    for q in rows:
        ws.append([
            q["origin"], q["destination"], q.get("hub") or "",
            q["outbound_date"], q.get("return_date") or "",
            float(q["price"]),
            q.get("airline") or "",
            round((q.get("duration_min") or 0) / 60, 1),
            q.get("transfers") or 0,
            _to_brt(q.get("fetched_at")),
        ])

    if len(rows) >= 2:
        last = ws.max_row
        rule = ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
        ws.conditional_formatting.add(f"F2:F{last}", rule)

    _autosize_columns(ws)


def _sheet_history(wb, storage, config: dict):
    """Serie temporal: para cada hub, lista preco minimo por sweep."""
    ws = wb.create_sheet("Histórico")
    ws.append(["Hub", "Timestamp (BRT)", "Preço mínimo (BRL)"])
    _header_style(ws, 1)

    trip = config["trip"]
    history = storage.get_history_by_hub(
        trip["origin"], trip["destination"],
        trip["outbound_date"], trip["return_date"],
    )
    for hub, points in sorted(history.items()):
        for p in points:
            ws.append([hub, _to_brt(p["ts"]), float(p["price"])])

    _autosize_columns(ws)


def _sheet_divergences(wb, storage, config: dict):
    """Quando ha cotacoes serpapi (verificacao top-N), compara com travelpayouts/principal."""
    ws = wb.create_sheet("Divergências")
    ws.append([
        "Rota", "Hub", "Datas", "Principal (BRL)", "Verificação (BRL)",
        "Diferença (BRL)", "Diferença %", "Vencedor",
    ])
    _header_style(ws, 1)

    trip = config["trip"]
    tp_latest = {(q.get("hub") or ""): q
                 for q in storage.get_latest_by_hub(
                     trip["origin"], trip["destination"],
                     trip["outbound_date"], trip["return_date"], source="travelpayouts")}
    sp_latest = {(q.get("hub") or ""): q
                 for q in storage.get_latest_by_hub(
                     trip["origin"], trip["destination"],
                     trip["outbound_date"], trip["return_date"], source="serpapi")}

    for hub, sp_q in sp_latest.items():
        tp_q = tp_latest.get(hub)
        if not tp_q:
            continue
        diff = sp_q["price"] - tp_q["price"]
        pct = (diff / tp_q["price"] * 100) if tp_q["price"] else 0
        winner = "Principal" if tp_q["price"] < sp_q["price"] else "Verificação"
        ws.append([
            f'{trip["origin"]}↔{trip["destination"]}',
            hub,
            f'{trip["outbound_date"]} / {trip["return_date"]}',
            float(tp_q["price"]), float(sp_q["price"]),
            round(float(diff), 2), round(float(pct), 1), winner,
        ])

    _autosize_columns(ws)


def _header_style(ws, row: int):
    fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = f"A{row + 1}"


def _autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


# ---------- HTML ----------

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="UTF-8">
<title>Monitor de Voos {{origin}}{{trip_arrow}}{{destination}}</title>
<style>
  * { box-sizing: border-box; }
  body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
         margin: 0; padding: 24px; background: #f5f7fa; color: #1f2937; }
  h1 { margin: 0 0 4px; color: #0f172a; }
  .subtitle { color: #64748b; margin-bottom: 24px; font-size: 14px; }
  .grid { display: grid; gap: 16px; }
  .card { background: white; border-radius: 8px; padding: 20px;
          box-shadow: 0 1px 3px rgba(0,0,0,0.06); }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { background: #1e3a8a; color: white; padding: 10px 8px; text-align: left;
       position: sticky; top: 0; }
  td { padding: 8px; border-bottom: 1px solid #e5e7eb; }
  tr:hover { background: #f9fafb; }
  .price { font-weight: 600; }
  .price-best { color: #16a34a; }
  .price-mid { color: #d97706; }
  .price-worst { color: #dc2626; }
  .badge { display: inline-block; padding: 2px 8px; border-radius: 4px;
           font-size: 11px; background: #e0e7ff; color: #3730a3; font-weight: 600; }
  .badge-direct { background: #dcfce7; color: #166534; }
  .badge-window { background: #fef3c7; color: #92400e; }
  tr.in-window td { background: #fffbeb; }
  tr.in-window:hover td { background: #fef3c7; }
  .window-info { padding: 8px 12px; background: #fef3c7; border-left: 4px solid #f59e0b;
                 border-radius: 4px; margin-bottom: 12px; font-size: 13px; color: #92400e; }
  .chart-wrap { height: 360px; }
  .footer { text-align: center; color: #9ca3af; font-size: 12px; margin-top: 24px; }
  .stat-row { display: flex; gap: 16px; flex-wrap: wrap; }
  .stat { flex: 1; min-width: 160px; }
  .stat-label { color: #64748b; font-size: 12px; text-transform: uppercase; letter-spacing: 0.5px; }
  .stat-value { font-size: 24px; font-weight: 700; color: #0f172a; margin-top: 4px; }
</style>
</head>
<body>
<h1>Monitor de Voos &mdash; {{origin}} {{trip_arrow}} {{destination}} {{destination_name}}</h1>
<div class="subtitle">
  Ida: <strong>{{outbound_date}}</strong>{{return_block}}
  &middot; {{passengers}} adulto(s), Econ&ocirc;mica, com bagagem despachada &middot; Atualizado em {{generated_at}}
</div>

<div class="grid">

  <div class="card">
    <div class="stat-row">
      <div class="stat">
        <div class="stat-label">Melhor preço atual</div>
        <div class="stat-value">{{best_total}}</div>
        <div style="color:#64748b; font-size:13px;">via {{best_hub}}</div>
      </div>
      <div class="stat">
        <div class="stat-label">Mínimo histórico (melhor hub)</div>
        <div class="stat-value">{{historical_min}}</div>
      </div>
      <div class="stat">
        <div class="stat-label">Hubs cotados</div>
        <div class="stat-value">{{hubs_with_data}}</div>
      </div>
      <div class="stat">
        <div class="stat-label">Total cotações no banco</div>
        <div class="stat-value">{{total_quotes}}</div>
      </div>
    </div>
  </div>

  <div class="card">
    <h2 style="margin-top:0;">Top opções por hub de conexão ({{trip_kind}} {{origin}} {{trip_arrow}} {{destination}})</h2>
    {{window_banner}}
    <table>
      <thead><tr>
        <th>#</th><th>Hub</th><th>Cidade</th>
        <th>Preço total</th><th>Companhia(s)</th>
        <th>Duração</th><th>Escalas</th><th>Chegada {{destination}}</th><th>Atualizado</th>
      </tr></thead>
      <tbody>
        {{hub_rows}}
      </tbody>
    </table>
  </div>

  <div class="card">
    <h2 style="margin-top:0;">Evolução do preço por hub</h2>
    <div class="chart-wrap"><canvas id="chart"></canvas></div>
  </div>

</div>

<div class="footer">
  Cotações do Google Flights via SerpAPI.
  <strong>Sempre confira no Google Flights antes de comprar.</strong>
</div>

<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0"></script>
<script>
const data = {{chart_data_json}};
const ctx = document.getElementById('chart').getContext('2d');
new Chart(ctx, {
  type: 'line',
  data: data,
  options: {
    responsive: true,
    maintainAspectRatio: false,
    interaction: { mode: 'index', intersect: false },
    scales: {
      y: {
        title: { display: true, text: 'Preço (BRL)' },
        ticks: { callback: v => 'R$ ' + v.toLocaleString('pt-BR') }
      },
      x: { title: { display: true, text: 'Cotação (horário BRT)' } }
    },
    plugins: {
      legend: { position: 'bottom' },
      tooltip: { callbacks: {
        label: ctx => ctx.dataset.label + ': R$ ' + ctx.parsed.y.toLocaleString('pt-BR')
      }}
    }
  }
});
</script>
</body>
</html>
"""


def _fmt_brl(n: float) -> str:
    s = f"R$ {n:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_duration(minutes: int) -> str:
    if not minutes:
        return "—"
    h, m = divmod(int(minutes), 60)
    return f"{h}h{m:02d}"


def generate_html(storage, config: dict, output_path: str):
    trip = config["trip"]
    hubs_by_code = {h["code"]: h for h in config["hubs"]}
    dest_name = trip.get("destination_name", "")
    dest_name_str = f"({dest_name})" if dest_name else ""
    arrival_window = (config.get("monitoring") or {}).get("arrival_window")
    window_active = has_window(config)

    return_date = trip.get("return_date") or None
    is_one_way = not return_date
    trip_kind = "one-way" if is_one_way else "round-trip"
    trip_arrow = "&rarr;" if is_one_way else "&harr;"
    return_block = "" if is_one_way else f" &middot; Volta: <strong>{return_date}</strong>"

    if window_active:
        window_banner = (
            f'<div class="window-info">Voos destacados em amarelo tem chegada em '
            f'<strong>{trip["destination"]}</strong> entre <strong>{window_label(config)}</strong> '
            f'(horario local). Alertas por email so disparam para esses voos.</div>'
        )
    else:
        window_banner = ""

    rows = storage.get_latest_by_hub(
        trip["origin"], trip["destination"],
        trip["outbound_date"], trip["return_date"],
    )

    hub_rows_html = ""
    for i, q in enumerate(rows, 1):
        hub = q.get("hub") or "—"
        city = hubs_by_code.get(hub, {}).get("city", "")
        cls = "price-best" if i <= 3 else ("price-mid" if i <= 7 else "price-worst")
        is_direct = (q.get("transfers") or 0) == 0
        badge_cls = "badge badge-direct" if is_direct else "badge"
        arrival_str = q.get("arrival_time") or ""
        in_window = window_active and in_arrival_window(arrival_str, arrival_window)
        tr_cls = ' class="in-window"' if in_window else ""
        # Mostra so HH:MM (hora local do destino) na coluna de chegada
        arr_display = "—"
        if arrival_str:
            t = str(arrival_str)
            if " " in t:
                t = t.split(" ")[-1]
            if "T" in t:
                t = t.split("T")[-1]
            arr_display = t[:5] if len(t) >= 5 else t
        arrival_cell = (
            f'<span class="badge badge-window">{arr_display}</span>'
            if in_window
            else arr_display
        )
        hub_rows_html += (
            f'<tr{tr_cls}>'
            f'<td>{i}</td>'
            f'<td><span class="{badge_cls}">{hub}</span></td>'
            f'<td>{city or "—"}</td>'
            f'<td class="price {cls}">{_fmt_brl(q["price"])}</td>'
            f'<td>{q.get("airline") or "—"}</td>'
            f'<td>{_fmt_duration(q.get("duration_min"))}</td>'
            f'<td>{q.get("transfers") or 0}</td>'
            f'<td>{arrival_cell}</td>'
            f'<td>{_to_brt(q.get("fetched_at"))}</td>'
            f'</tr>'
        )

    if rows:
        best = rows[0]
        best_total_str = _fmt_brl(best["price"])
        best_hub_str = best.get("hub") or "—"
        # Min historico do hub vencedor
        hist_min = storage.get_historical_min_by_hub(
            trip["origin"], trip["destination"],
            trip["outbound_date"], trip["return_date"],
            best_hub_str,
        )
        historical_min_str = _fmt_brl(hist_min) if hist_min else "—"
    else:
        best_total_str = "—"
        best_hub_str = "—"
        historical_min_str = "—"

    hubs_with_data = len(rows)
    total_quotes = _total_quotes(storage)
    chart_data = _build_chart_data(storage, config, hubs_by_code)

    html = (HTML_TEMPLATE
        .replace("{{origin}}", trip["origin"])
        .replace("{{destination}}", trip["destination"])
        .replace("{{destination_name}}", dest_name_str)
        .replace("{{outbound_date}}", trip["outbound_date"])
        .replace("{{return_block}}", return_block)
        .replace("{{trip_kind}}", trip_kind)
        .replace("{{trip_arrow}}", trip_arrow)
        .replace("{{window_banner}}", window_banner)
        .replace("{{passengers}}", str(trip["passengers"]))
        .replace("{{generated_at}}", _to_brt(datetime.utcnow().isoformat(timespec="seconds")))
        .replace("{{best_total}}", best_total_str)
        .replace("{{best_hub}}", best_hub_str)
        .replace("{{historical_min}}", historical_min_str)
        .replace("{{hubs_with_data}}", str(hubs_with_data))
        .replace("{{total_quotes}}", str(total_quotes))
        .replace("{{hub_rows}}", hub_rows_html)
        .replace("{{chart_data_json}}", json.dumps(chart_data))
    )

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    log.info(f"  HTML salvo: {output_path}")


def _total_quotes(storage) -> int:
    with sqlite3.connect(storage.db_path) as c:
        return c.execute("SELECT COUNT(*) FROM quotes").fetchone()[0]


def _build_chart_data(storage, config: dict, hubs_by_code: dict):
    """Para cada hub, serie de pontos (dia, preco min)."""
    trip = config["trip"]
    history = storage.get_history_by_hub(
        trip["origin"], trip["destination"],
        trip["outbound_date"], trip["return_date"],
    )

    all_ts: set[str] = set()
    for points in history.values():
        for p in points:
            all_ts.add(p["ts"])
    timestamps = sorted(all_ts)
    # Labels exibidos no grafico em BRT (DD/MM HH:MM), ordem mantida pelo UTC
    labels_brt = [_to_brt(t, fmt="%d/%m %H:%M") for t in timestamps]

    colors = ["#2563eb", "#dc2626", "#16a34a", "#9333ea", "#ea580c",
              "#0891b2", "#be123c", "#65a30d", "#7c3aed", "#c2410c",
              "#1e40af", "#a21caf", "#475569"]

    datasets = []
    # Ordena hubs por menor preco visto (mais baratos primeiro = primeiras cores)
    hubs_sorted = sorted(history.keys(),
                         key=lambda h: min(p["price"] for p in history[h]))
    for i, hub in enumerate(hubs_sorted):
        by_ts = {p["ts"]: p["price"] for p in history[hub]}
        data_points = [by_ts.get(t) for t in timestamps]
        if any(p is not None for p in data_points):
            city = hubs_by_code.get(hub, {}).get("city", "")
            label = f"{hub} ({city})" if city else hub
            datasets.append({
                "label": label,
                "data": data_points,
                "borderColor": colors[i % len(colors)],
                "backgroundColor": colors[i % len(colors)] + "33",
                "tension": 0.2,
                "spanGaps": True,
            })

    return {"labels": labels_brt, "datasets": datasets}
